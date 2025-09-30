import os
import requests
import concurrent.futures
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pdf2image import convert_from_path
import logging
import shutil

# Set up logging
logging.basicConfig(level=logging.INFO)

def process_excel_hyperlinks(
    excel_path: str,
    output_path: str,
    temp_dir: str = "temp_files",
    poppler_path: str = r"C:\Program Files\poppler\bin",
    max_workers: int = 4
) -> None:
    """Optimized Excel hyperlink processor."""
    os.makedirs(temp_dir, exist_ok=True)
    _validate_poppler(poppler_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    video_exts = {'mp4', 'avi', 'mov', 'mkv', 'wmv'}

    # Group hyperlinks by row
    row_hyperlinks = _group_hyperlinks(ws)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for row, links in row_hyperlinks.items():
            links.sort(key=lambda x: x["col"])  # Sort links by column
            start_col = _find_insert_start_col(ws, row, len(links))  # Calculate starting column
            
            for idx, link in enumerate(links):
                futures.append(
                    executor.submit(
                        _process_single_link,
                        link, row, start_col + idx,
                        temp_dir, poppler_path, video_exts, ws
                    )
                )

        concurrent.futures.wait(futures)

    # Adjust column widths and row heights
    _adjust_dimensions(ws)

    wb.save(output_path)
    shutil.rmtree(temp_dir)  # Clean up temporary files

def _group_hyperlinks(ws) -> dict:
    """Group hyperlinks by row."""
    row_hyperlinks = {}
    for cell in ws.iter_rows():
        for c in cell:
            if c.hyperlink:
                row = c.row
                if row not in row_hyperlinks:
                    row_hyperlinks[row] = []
                row_hyperlinks[row].append({
                    "cell": c,
                    "url": c.hyperlink.target,
                    "col": c.column
                })
    return row_hyperlinks

def _find_insert_start_col(ws, row, link_count) -> int:
    """Calculate the starting column for inserting images."""
    max_hyperlink_col = 0
    if link_count > 0:
        for cell in ws[row]:
            if cell.hyperlink:
                max_hyperlink_col = max(max_hyperlink_col, cell.column)
    return max_hyperlink_col + 1

def _validate_poppler(poppler_path: str) -> None:
    """Validate Poppler path."""
    required_files = {'pdftoppm.exe', 'pdfinfo.exe'}
    poppler_bin = os.listdir(poppler_path)
    missing_files = required_files - set(poppler_bin)
    if missing_files:
        raise FileNotFoundError(f"Missing critical Poppler files: {missing_files}")

def _process_single_link(link: dict, row: int, start_col: int, temp_dir: str, 
                        poppler_path: str, video_exts: set, ws) -> None:
    """Process a single hyperlink."""
    try:
        response = requests.get(link["url"], timeout=15, stream=True)
        response.raise_for_status()
        ext = _get_file_extension(link["url"])
        
        if not ext:
            logging.warning(f"Unrecognized file type: {link['url']} (Row: {row}, Col: {link['col']})")
            return
        if ext in video_exts:
            logging.info(f"Skipping video file: {link['url']} (Row: {row}, Col: {link['col']})")
            return

        file_path = os.path.join(temp_dir, f"row{row}_col{link['col']}.{ext}")
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        # Handle PDF files
        if ext == "pdf":
            images = convert_from_path(file_path, poppler_path=poppler_path)
            for page_num, img in enumerate(images, start=1):
                img_path = os.path.join(temp_dir, f"row{row}_col{link['col']}_p{page_num}.jpg")
                img.save(img_path, "JPEG")
                _insert_image(img_path, row, start_col + page_num - 1, ws)
        # Handle image files
        elif ext in {"jpg", "jpeg", "png"}:
            _insert_image(file_path, row, start_col, ws)
        else:
            logging.warning(f"Unsupported file type: {ext} (Row: {row}, Col: {link['col']})")

    except requests.exceptions.RequestException as e:
        logging.error(f"Download failed: {link['url']} - {str(e)} (Row: {row}, Col: {link['col']})")
    except Exception as e:
        logging.error(f"Processing error: {link['url']} - {str(e)} (Row: {row}, Col: {link['col']})")

def _get_file_extension(url: str) -> str:
    """Get the file extension from a URL."""
    parsed_path = urlparse(url).path
    return os.path.splitext(parsed_path)[1].lstrip('.').lower()

def _insert_image(img_path: str, row: int, col: int, ws) -> None:
    """Insert an image into the worksheet."""
    img = Image(img_path)
    img.width, img.height = 180, 120  # Maintain aspect ratio
    cell_ref = f"{get_column_letter(col)}{row}"
    ws.add_image(img, cell_ref)

def _adjust_dimensions(ws) -> None:
    """Adjust column widths and row heights."""
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 120

if __name__ == "__main__":
    process_excel_hyperlinks(
        excel_path="D:\\python\\test1.xlsx",
        output_path="D:\\python\\1.xlsx",
        temp_dir="D:\\python\\downloaded_files",
        poppler_path=r"C:\Program Files\poppler\bin"
    )
    
    input("Press Enter to exit...")  # Wait for user input before closing